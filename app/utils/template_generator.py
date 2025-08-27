import json
import csv
import io
from datetime import datetime, date
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

from app.models.data_import import ImportType, FileType, ImportTemplate

class TemplateGenerator:
    """Generate import templates for different data types"""
    
    def __init__(self):
        self.templates = {
            ImportType.PRODUCTS: self._get_product_template,
            ImportType.HISTORICAL_SALES: self._get_sales_template,
            ImportType.INVENTORY_LEVELS: self._get_inventory_template,
            ImportType.MANUFACTURING_DATA: self._get_manufacturing_template,
            ImportType.FINANCIAL_DATA: self._get_financial_template,
            ImportType.FORECASTS: self._get_forecast_template,
        }
    
    def generate_template(self, import_type: ImportType, file_format: FileType) -> Dict[str, Any]:
        """Generate template for specific import type and format"""
        if import_type not in self.templates:
            raise ValueError(f"Template not available for import type: {import_type.value}")
        
        template_data = self.templates[import_type]()
        
        if file_format == FileType.CSV:
            return self._generate_csv_template(template_data)
        elif file_format == FileType.XLSX:
            return self._generate_excel_template(template_data)
        elif file_format == FileType.JSON:
            return self._generate_json_template(template_data)
        else:
            raise ValueError(f"Unsupported file format: {file_format.value}")
    
    def create_system_templates(self) -> List[ImportTemplate]:
        """Create system default templates"""
        templates = []
        
        for import_type in ImportType:
            for file_format in [FileType.CSV, FileType.XLSX, FileType.JSON]:
                try:
                    template_data = self.templates[import_type]()
                    
                    template = ImportTemplate(
                        template_name=f"{import_type.value.title()} Template ({file_format.value.upper()})",
                        import_type=import_type,
                        version="1.0",
                        description=f"Standard template for importing {import_type.value} data",
                        file_format=file_format,
                        field_definitions=template_data['field_definitions'],
                        sample_data=template_data['sample_data'],
                        validation_rules=template_data['validation_rules'],
                        is_system_template=True
                    )
                    
                    templates.append(template)
                    
                except KeyError:
                    # Skip if template not defined for this type
                    continue
        
        return templates
    
    def _get_product_template(self) -> Dict[str, Any]:
        """Get product template definition"""
        return {
            'field_definitions': {
                'sku': {
                    'type': 'string',
                    'required': True,
                    'max_length': 50,
                    'pattern': r'^[A-Z0-9\-_]{3,50}$',
                    'description': 'Unique product SKU (format: PRODUCT-REGION-VARIANT)',
                    'example': 'GABA-RED-UK-001'
                },
                'name': {
                    'type': 'string',
                    'required': True,
                    'max_length': 100,
                    'description': 'Product name',
                    'example': 'GABA Red - UK Formula'
                },
                'category': {
                    'type': 'string',
                    'required': True,
                    'allowed_values': ['GABA Red', 'GABA Black', 'GABA Gold'],
                    'description': 'Product category',
                    'example': 'GABA Red'
                },
                'market_region': {
                    'type': 'string',
                    'required': True,
                    'allowed_values': ['UK', 'EU', 'USA'],
                    'description': 'Target market region',
                    'example': 'UK'
                },
                'weight_kg': {
                    'type': 'decimal',
                    'required': False,
                    'min_value': 0.001,
                    'max_value': 50.000,
                    'decimal_places': 3,
                    'description': 'Product weight in kilograms',
                    'example': 0.250
                },
                'dimensions_cm': {
                    'type': 'string',
                    'required': False,
                    'pattern': r'^\d+(\.\d+)?x\d+(\.\d+)?x\d+(\.\d+)?$',
                    'description': 'Product dimensions in cm (LxWxH)',
                    'example': '10.5x5.2x15.0'
                },
                'unit_cost': {
                    'type': 'decimal',
                    'required': False,
                    'min_value': 0.01,
                    'max_value': 10000.00,
                    'decimal_places': 2,
                    'description': 'Unit cost in local currency',
                    'example': 12.50
                },
                'selling_price': {
                    'type': 'decimal',
                    'required': False,
                    'min_value': 0.01,
                    'max_value': 10000.00,
                    'decimal_places': 2,
                    'description': 'Selling price in local currency',
                    'example': 24.99
                },
                'production_time_hours': {
                    'type': 'decimal',
                    'required': False,
                    'min_value': 0.1,
                    'max_value': 72.0,
                    'decimal_places': 2,
                    'description': 'Production time in hours',
                    'example': 2.5
                },
                'batch_size_min': {
                    'type': 'integer',
                    'required': False,
                    'min_value': 1,
                    'max_value': 10000,
                    'description': 'Minimum batch size',
                    'example': 100
                },
                'batch_size_max': {
                    'type': 'integer',
                    'required': False,
                    'min_value': 1,
                    'max_value': 100000,
                    'description': 'Maximum batch size',
                    'example': 5000
                }
            },
            'sample_data': [
                {
                    'sku': 'GABA-RED-UK-001',
                    'name': 'GABA Red - UK Formula',
                    'category': 'GABA Red',
                    'market_region': 'UK',
                    'weight_kg': 0.250,
                    'dimensions_cm': '10.5x5.2x15.0',
                    'unit_cost': 12.50,
                    'selling_price': 24.99,
                    'production_time_hours': 2.5,
                    'batch_size_min': 100,
                    'batch_size_max': 5000
                },
                {
                    'sku': 'GABA-BLACK-EU-001',
                    'name': 'GABA Black - EU Formula',
                    'category': 'GABA Black',
                    'market_region': 'EU',
                    'weight_kg': 0.300,
                    'dimensions_cm': '11.0x5.5x16.0',
                    'unit_cost': 18.75,
                    'selling_price': 39.99,
                    'production_time_hours': 3.0,
                    'batch_size_min': 50,
                    'batch_size_max': 3000
                }
            ],
            'validation_rules': {
                'business_rules': [
                    'selling_price must be greater than unit_cost',
                    'batch_size_max must be greater than batch_size_min',
                    'SKU must be unique across all products'
                ],
                'cross_field_validation': [
                    'If unit_cost is provided, selling_price is required',
                    'If batch_size_min is provided, batch_size_max is required'
                ]
            }
        }
    
    def _get_sales_template(self) -> Dict[str, Any]:
        """Get historical sales template definition"""
        return {
            'field_definitions': {
                'product_sku': {
                    'type': 'string',
                    'required': True,
                    'description': 'Product SKU (must exist in product catalog)',
                    'example': 'GABA-RED-UK-001'
                },
                'sales_channel': {
                    'type': 'string',
                    'required': True,
                    'allowed_values': ['Amazon UK', 'Amazon USA', 'Shopify UK', 'Shopify EU', 'Shopify USA'],
                    'description': 'Sales channel name',
                    'example': 'Amazon UK'
                },
                'sale_date': {
                    'type': 'date',
                    'required': True,
                    'format': 'YYYY-MM-DD',
                    'min_date': '2020-01-01',
                    'description': 'Date of sale',
                    'example': '2024-01-15'
                },
                'sale_time': {
                    'type': 'time',
                    'required': False,
                    'format': 'HH:MM:SS',
                    'description': 'Time of sale (optional)',
                    'example': '14:30:00'
                },
                'quantity_sold': {
                    'type': 'integer',
                    'required': True,
                    'min_value': 1,
                    'max_value': 10000,
                    'description': 'Quantity sold',
                    'example': 2
                },
                'unit_price': {
                    'type': 'decimal',
                    'required': True,
                    'min_value': 0.01,
                    'max_value': 1000.00,
                    'decimal_places': 2,
                    'description': 'Unit price in local currency',
                    'example': 24.99
                },
                'currency': {
                    'type': 'string',
                    'required': True,
                    'allowed_values': ['GBP', 'EUR', 'USD'],
                    'description': 'Currency code',
                    'example': 'GBP'
                },
                'gross_revenue': {
                    'type': 'decimal',
                    'required': True,
                    'decimal_places': 2,
                    'description': 'Gross revenue (quantity × unit_price)',
                    'example': 49.98
                },
                'discounts': {
                    'type': 'decimal',
                    'required': False,
                    'min_value': 0.00,
                    'decimal_places': 2,
                    'description': 'Total discounts applied',
                    'example': 5.00
                },
                'shipping_cost': {
                    'type': 'decimal',
                    'required': False,
                    'min_value': 0.00,
                    'decimal_places': 2,
                    'description': 'Shipping cost charged to customer',
                    'example': 3.99
                },
                'platform_fees': {
                    'type': 'decimal',
                    'required': False,
                    'min_value': 0.00,
                    'decimal_places': 2,
                    'description': 'Platform fees deducted',
                    'example': 7.50
                },
                'taxes': {
                    'type': 'decimal',
                    'required': False,
                    'min_value': 0.00,
                    'decimal_places': 2,
                    'description': 'Taxes collected/paid',
                    'example': 8.33
                },
                'order_id': {
                    'type': 'string',
                    'required': False,
                    'max_length': 100,
                    'description': 'Platform order ID',
                    'example': 'AMZ-12345-67890'
                },
                'customer_type': {
                    'type': 'string',
                    'required': False,
                    'allowed_values': ['B2C', 'B2B'],
                    'description': 'Customer type',
                    'example': 'B2C'
                },
                'fulfillment_method': {
                    'type': 'string',
                    'required': False,
                    'allowed_values': ['FBA', 'FBM', 'Own'],
                    'description': 'Fulfillment method',
                    'example': 'FBA'
                },
                'shipping_country': {
                    'type': 'string',
                    'required': False,
                    'description': 'Shipping country code',
                    'example': 'GB'
                }
            },
            'sample_data': [
                {
                    'product_sku': 'GABA-RED-UK-001',
                    'sales_channel': 'Amazon UK',
                    'sale_date': '2024-01-15',
                    'sale_time': '14:30:00',
                    'quantity_sold': 2,
                    'unit_price': 24.99,
                    'currency': 'GBP',
                    'gross_revenue': 49.98,
                    'discounts': 5.00,
                    'shipping_cost': 0.00,
                    'platform_fees': 7.50,
                    'taxes': 8.33,
                    'order_id': 'AMZ-UK-12345',
                    'customer_type': 'B2C',
                    'fulfillment_method': 'FBA',
                    'shipping_country': 'GB'
                }
            ],
            'validation_rules': {
                'business_rules': [
                    'gross_revenue must equal quantity_sold × unit_price',
                    'discounts cannot exceed gross_revenue',
                    'sale_date cannot be in the future',
                    'product_sku must exist in product catalog'
                ],
                'currency_rules': [
                    'GBP for UK sales channels',
                    'EUR for EU sales channels',
                    'USD for USA sales channels'
                ]
            }
        }
    
    def _get_inventory_template(self) -> Dict[str, Any]:
        """Get inventory levels template definition"""
        return {
            'field_definitions': {
                'product_sku': {
                    'type': 'string',
                    'required': True,
                    'description': 'Product SKU',
                    'example': 'GABA-RED-UK-001'
                },
                'location': {
                    'type': 'string',
                    'required': True,
                    'description': 'Inventory location',
                    'example': 'UK Warehouse'
                },
                'available_quantity': {
                    'type': 'integer',
                    'required': True,
                    'min_value': 0,
                    'description': 'Available inventory quantity',
                    'example': 1500
                },
                'reserved_quantity': {
                    'type': 'integer',
                    'required': False,
                    'min_value': 0,
                    'description': 'Reserved inventory quantity',
                    'example': 250
                },
                'reorder_point': {
                    'type': 'integer',
                    'required': False,
                    'min_value': 0,
                    'description': 'Reorder point threshold',
                    'example': 500
                },
                'max_stock_level': {
                    'type': 'integer',
                    'required': False,
                    'min_value': 0,
                    'description': 'Maximum stock level',
                    'example': 5000
                },
                'last_updated': {
                    'type': 'datetime',
                    'required': True,
                    'format': 'YYYY-MM-DD HH:MM:SS',
                    'description': 'Last update timestamp',
                    'example': '2024-01-15 10:30:00'
                }
            },
            'sample_data': [
                {
                    'product_sku': 'GABA-RED-UK-001',
                    'location': 'UK Warehouse',
                    'available_quantity': 1500,
                    'reserved_quantity': 250,
                    'reorder_point': 500,
                    'max_stock_level': 5000,
                    'last_updated': '2024-01-15 10:30:00'
                }
            ],
            'validation_rules': {
                'business_rules': [
                    'available_quantity + reserved_quantity <= max_stock_level',
                    'reorder_point < max_stock_level',
                    'last_updated cannot be in the future'
                ]
            }
        }
    
    def _get_manufacturing_template(self) -> Dict[str, Any]:
        """Get manufacturing data template definition"""
        return {
            'field_definitions': {
                'batch_number': {
                    'type': 'string',
                    'required': True,
                    'description': 'Unique batch identifier',
                    'example': 'BATCH-2024-001'
                },
                'product_sku': {
                    'type': 'string',
                    'required': True,
                    'description': 'Product SKU being manufactured',
                    'example': 'GABA-RED-UK-001'
                },
                'production_date': {
                    'type': 'date',
                    'required': True,
                    'format': 'YYYY-MM-DD',
                    'description': 'Production date',
                    'example': '2024-01-15'
                },
                'planned_quantity': {
                    'type': 'integer',
                    'required': True,
                    'min_value': 1,
                    'description': 'Planned production quantity',
                    'example': 1000
                },
                'actual_quantity': {
                    'type': 'integer',
                    'required': False,
                    'min_value': 0,
                    'description': 'Actual production quantity',
                    'example': 985
                },
                'yield_percentage': {
                    'type': 'decimal',
                    'required': False,
                    'min_value': 0.0,
                    'max_value': 100.0,
                    'decimal_places': 2,
                    'description': 'Production yield percentage',
                    'example': 98.5
                },
                'quality_score': {
                    'type': 'decimal',
                    'required': False,
                    'min_value': 0.0,
                    'max_value': 100.0,
                    'decimal_places': 2,
                    'description': 'Quality control score',
                    'example': 95.5
                },
                'production_cost': {
                    'type': 'decimal',
                    'required': False,
                    'min_value': 0.01,
                    'decimal_places': 2,
                    'description': 'Total production cost',
                    'example': 12500.00
                }
            },
            'sample_data': [
                {
                    'batch_number': 'BATCH-2024-001',
                    'product_sku': 'GABA-RED-UK-001',
                    'production_date': '2024-01-15',
                    'planned_quantity': 1000,
                    'actual_quantity': 985,
                    'yield_percentage': 98.5,
                    'quality_score': 95.5,
                    'production_cost': 12500.00
                }
            ],
            'validation_rules': {
                'business_rules': [
                    'actual_quantity should be <= planned_quantity',
                    'yield_percentage = (actual_quantity / planned_quantity) × 100',
                    'quality_score should be >= 70% for acceptable quality'
                ]
            }
        }
    
    def _get_financial_template(self) -> Dict[str, Any]:
        """Get financial data template definition"""
        return {
            'field_definitions': {
                'transaction_date': {
                    'type': 'date',
                    'required': True,
                    'format': 'YYYY-MM-DD',
                    'description': 'Transaction date',
                    'example': '2024-01-15'
                },
                'transaction_type': {
                    'type': 'string',
                    'required': True,
                    'allowed_values': ['Revenue', 'Cost', 'Investment', 'Tax', 'Fee'],
                    'description': 'Type of transaction',
                    'example': 'Revenue'
                },
                'amount': {
                    'type': 'decimal',
                    'required': True,
                    'decimal_places': 2,
                    'description': 'Transaction amount',
                    'example': 1500.00
                },
                'currency': {
                    'type': 'string',
                    'required': True,
                    'allowed_values': ['GBP', 'EUR', 'USD'],
                    'description': 'Currency code',
                    'example': 'GBP'
                },
                'account_code': {
                    'type': 'string',
                    'required': False,
                    'description': 'Account code',
                    'example': 'REV-001'
                },
                'description': {
                    'type': 'string',
                    'required': True,
                    'max_length': 255,
                    'description': 'Transaction description',
                    'example': 'Amazon UK sales revenue'
                }
            },
            'sample_data': [
                {
                    'transaction_date': '2024-01-15',
                    'transaction_type': 'Revenue',
                    'amount': 1500.00,
                    'currency': 'GBP',
                    'account_code': 'REV-001',
                    'description': 'Amazon UK sales revenue'
                }
            ],
            'validation_rules': {
                'business_rules': [
                    'Amount must be positive for Revenue and Investment',
                    'Amount must be negative for Cost, Tax, and Fee'
                ]
            }
        }
    
    def _get_forecast_template(self) -> Dict[str, Any]:
        """Get forecast data template definition"""
        return {
            'field_definitions': {
                'product_sku': {
                    'type': 'string',
                    'required': True,
                    'description': 'Product SKU',
                    'example': 'GABA-RED-UK-001'
                },
                'forecast_period': {
                    'type': 'string',
                    'required': True,
                    'pattern': r'^\d{4}-\d{2}$',
                    'description': 'Forecast period (YYYY-MM)',
                    'example': '2024-02'
                },
                'sales_channel': {
                    'type': 'string',
                    'required': True,
                    'description': 'Sales channel',
                    'example': 'Amazon UK'
                },
                'forecasted_quantity': {
                    'type': 'integer',
                    'required': True,
                    'min_value': 0,
                    'description': 'Forecasted sales quantity',
                    'example': 500
                },
                'confidence_level': {
                    'type': 'decimal',
                    'required': False,
                    'min_value': 0.0,
                    'max_value': 100.0,
                    'decimal_places': 2,
                    'description': 'Forecast confidence level (%)',
                    'example': 85.5
                },
                'forecast_method': {
                    'type': 'string',
                    'required': False,
                    'allowed_values': ['Historical Average', 'Trend Analysis', 'ML Model', 'Manual'],
                    'description': 'Forecasting method used',
                    'example': 'ML Model'
                }
            },
            'sample_data': [
                {
                    'product_sku': 'GABA-RED-UK-001',
                    'forecast_period': '2024-02',
                    'sales_channel': 'Amazon UK',
                    'forecasted_quantity': 500,
                    'confidence_level': 85.5,
                    'forecast_method': 'ML Model'
                }
            ],
            'validation_rules': {
                'business_rules': [
                    'forecast_period must be in the future',
                    'confidence_level should be between 50% and 100%'
                ]
            }
        }
    
    def _generate_csv_template(self, template_data: Dict) -> Dict[str, Any]:
        """Generate CSV template"""
        field_definitions = template_data['field_definitions']
        sample_data = template_data['sample_data']
        
        # Create CSV content
        output = io.StringIO()
        fieldnames = list(field_definitions.keys())
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for row in sample_data:
            writer.writerow(row)
        
        csv_content = output.getvalue()
        output.close()
        
        return {
            'content': csv_content,
            'filename': 'template.csv',
            'content_type': 'text/csv'
        }
    
    def _generate_excel_template(self, template_data: Dict) -> Dict[str, Any]:
        """Generate Excel template with validation and formatting"""
        field_definitions = template_data['field_definitions']
        sample_data = template_data['sample_data']
        validation_rules = template_data.get('validation_rules', {})
        
        # Create workbook
        wb = Workbook()
        
        # Main data sheet
        ws_data = wb.active
        ws_data.title = "Data"
        
        # Instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        
        # Field definitions sheet
        ws_fields = wb.create_sheet("Field Definitions")
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        required_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Setup data sheet
        headers = list(field_definitions.keys())
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            
            # Mark required fields
            field_def = field_definitions[header]
            if field_def.get('required', False):
                cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            
            # Add comments with field descriptions
            if 'description' in field_def:
                comment_text = f"{field_def['description']}\n"
                if 'example' in field_def:
                    comment_text += f"Example: {field_def['example']}"
                
                comment = Comment(comment_text, "Template Generator")
                cell.comment = comment
        
        # Add sample data
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws_data.cell(row=row_num, column=col_num, value=row_data.get(header))
                cell.border = border
        
        # Add data validation
        self._add_excel_validation(ws_data, field_definitions, len(sample_data) + 10)
        
        # Setup instructions sheet
        self._create_instructions_sheet(ws_instructions, validation_rules)
        
        # Setup field definitions sheet
        self._create_field_definitions_sheet(ws_fields, field_definitions)
        
        # Auto-adjust column widths
        for ws in [ws_data, ws_fields]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            'content': output.getvalue(),
            'filename': 'template.xlsx',
            'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    
    def _add_excel_validation(self, ws, field_definitions: Dict, max_row: int):
        """Add data validation to Excel sheet"""
        for col_num, (field_name, field_def) in enumerate(field_definitions.items(), 1):
            col_letter = ws.cell(row=1, column=col_num).column_letter
            cell_range = f"{col_letter}2:{col_letter}{max_row}"
            
            # Add validation based on field type
            if 'allowed_values' in field_def:
                # Dropdown validation
                dv = DataValidation(type="list", formula1=f'"{",".join(field_def["allowed_values"])}"')
                dv.add(cell_range)
                ws.add_data_validation(dv)
            
            elif field_def.get('type') == 'integer':
                # Integer validation
                min_val = field_def.get('min_value', 0)
                max_val = field_def.get('max_value', 999999)
                dv = DataValidation(type="whole", operator="between", formula1=min_val, formula2=max_val)
                dv.add(cell_range)
                ws.add_data_validation(dv)
            
            elif field_def.get('type') == 'decimal':
                # Decimal validation
                min_val = field_def.get('min_value', 0)
                max_val = field_def.get('max_value', 999999)
                dv = DataValidation(type="decimal", operator="between", formula1=min_val, formula2=max_val)
                dv.add(cell_range)
                ws.add_data_validation(dv)
    
    def _create_instructions_sheet(self, ws, validation_rules: Dict):
        """Create instructions sheet"""
        ws.append(["Data Import Template Instructions"])
        ws.append([])
        ws.append(["1. Fill in the 'Data' sheet with your data"])
        ws.append(["2. Required fields are highlighted in red"])
        ws.append(["3. Hover over column headers to see field descriptions"])
        ws.append(["4. Follow the validation rules below"])
        ws.append([])
        
        if validation_rules.get('business_rules'):
            ws.append(["Business Rules:"])
            for rule in validation_rules['business_rules']:
                ws.append([f"• {rule}"])
            ws.append([])
        
        # Style the instructions
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith("•"):
                    cell.alignment = Alignment(indent=1)
    
    def _create_field_definitions_sheet(self, ws, field_definitions: Dict):
        """Create field definitions sheet"""
        # Headers
        headers = ["Field Name", "Type", "Required", "Description", "Constraints", "Example"]
        ws.append(headers)
        
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add field definitions
        for field_name, field_def in field_definitions.items():
            constraints = []
            
            if 'min_value' in field_def:
                constraints.append(f"Min: {field_def['min_value']}")
            if 'max_value' in field_def:
                constraints.append(f"Max: {field_def['max_value']}")
            if 'max_length' in field_def:
                constraints.append(f"Max length: {field_def['max_length']}")
            if 'allowed_values' in field_def:
                constraints.append(f"Values: {', '.join(field_def['allowed_values'])}")
            
            ws.append([
                field_name,
                field_def.get('type', 'string'),
                'Yes' if field_def.get('required', False) else 'No',
                field_def.get('description', ''),
                '; '.join(constraints),
                field_def.get('example', '')
            ])
    
    def _generate_json_template(self, template_data: Dict) -> Dict[str, Any]:
        """Generate JSON template"""
        template_structure = {
            "data": template_data['sample_data'],
            "_metadata": {
                "template_version": "1.0",
                "generated_at": datetime.now().isoformat(),
                "field_definitions": template_data['field_definitions'],
                "validation_rules": template_data.get('validation_rules', {}),
                "instructions": [
                    "Replace the sample data with your actual data",
                    "Ensure all required fields are provided",
                    "Follow the field definitions and validation rules",
                    "Remove this _metadata section before importing"
                ]
            }
        }
        
        json_content = json.dumps(template_structure, indent=2, default=str)
        
        return {
            'content': json_content,
            'filename': 'template.json',
            'content_type': 'application/json'
        }